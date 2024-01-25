from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# Excel Part
Excel = load_workbook("C:\\Users\\Arda\\Desktop\\ScriptForIMEItoSerial\\SampleExcel.xlsx")
ActivatedExcel = Excel.active

# Selenium Webdriver
driver = webdriver.Chrome()
url = "https://snlookup.com/samsung"
wait = WebDriverWait(driver, 10)

try:
    CurrentRow = 2
    for row in ActivatedExcel.iter_rows(min_row=2, max_col=1, values_only=True):
        for input_data in row:
            driver.get(url)

            search_box = driver.find_element("name", "sn")
            search_box.clear()
            search_box.send_keys(input_data)

            submit_button = driver.find_element("name", "search")
            submit_button.click()
        
            if str(input_data)[0] == 'R':
                ResultXPath =  '//*[@id="page-268"]/div[1]/table[1]/tbody/tr[5]/td[2]'  #The IMEI number is searched on the website
            else: 
                ResultXPath = '//*[@id="page-268"]/div[1]/table[1]/tbody/tr[6]/td[2]' #Serial number is searched on the website

            wait.until(EC.presence_of_all_elements_located((By.XPATH, ResultXPath)))
            loaded_element = driver.find_element(By.XPATH, ResultXPath)      
            ResultText = loaded_element.text
            print(f"Input_data = {input_data} - Output_data = {ResultText}")
            Excel.active.cell(row=CurrentRow, column=2, value=ResultText)
            CurrentRow += 1 

except Exception as e:
    print(f'Warning!: {e}')
    raise  

finally:
    driver.quit()
    Excel.save("C:\\Users\\Arda\\Desktop\\ScriptForIMEItoSerial\\SampleExcel.xlsx")
    print("Saved to Excel Sucess!")
